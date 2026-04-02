"""FILE SCANNER 핵심 엔진 함수 모음."""

from __future__ import annotations

import csv
import io
import logging
import os
import re
import threading
import warnings
import xml.etree.ElementTree as ET
import zipfile
import zlib
from datetime import datetime
from typing import Any, Callable, Iterable

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

try:
    import pdfplumber
except Exception:
    pdfplumber = None

try:
    import olefile
except Exception:
    olefile = None

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    import xlrd
except Exception:
    xlrd = None

try:
    from docx import Document
except Exception:
    Document = None


logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# 서드파티 라이브러리 경고 억제 (모듈 로드 시 1회)
# ──────────────────────────────────────────────
for _noisy in ("xlrd", "olefile", "openpyxl"):
    logging.getLogger(_noisy).setLevel(logging.CRITICAL)

warnings.filterwarnings("ignore", module=r"xlrd\..*")
warnings.filterwarnings("ignore", module=r"olefile\..*")


# ──────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────

DEFAULT_EXTENSIONS: set[str] = {
    ".pdf", ".xlsx", ".xls", ".docx", ".doc",
    ".hwp", ".hwpx", ".csv", ".txt",
}

EXCLUDED_DIR_NAMES: set[str] = {
    "windows", "program files", "program files (x86)",
    "$recycle.bin", "system volume information",
    "programdata", "appdata", "venv", "node_modules",
    ".git", "__pycache__", "dist", "build",
}

TXT_FALLBACK_ENCODINGS: tuple[str, ...] = ("utf-8", "cp949", "euc-kr", "mbcs")
QUICK_CHECK_ENCODINGS: tuple[str, ...] = ("utf-8", "cp949", "euc-kr", "utf-16-le")
QUICK_CHECK_MAX_BYTES: int = 10 * 1024 * 1024

BINARY_FALLBACK_MAX_BYTES: int = 10 * 1024 * 1024
MAX_EXTRACT_FILE_SIZE: int = 50 * 1024 * 1024
MAX_PDF_PAGES: int = 200
FILE_PROCESS_TIMEOUT: int = 15

MIN_WORKERS: int = 4
MAX_WORKERS: int = 8
CPU_DIVISOR: int = 1

_SEARCH_FILE_META: dict[str, tuple[bool, bool]] = {}
_SEARCH_FILE_META_LOCK = threading.Lock()


# ──────────────────────────────────────────────
# 파일 탐색
# ──────────────────────────────────────────────

def scan_files(
    paths: list[str],
    extensions: set[str] | None = DEFAULT_EXTENSIONS,
) -> list[str]:
    """지정 경로를 재귀 탐색해 대상 확장자 파일의 절대경로 목록을 반환한다."""
    if not paths:
        return []

    normalized_extensions: set[str] | None
    if extensions is None:
        normalized_extensions = None
    else:
        normalized_extensions = _normalize_extensions(extensions or DEFAULT_EXTENSIONS)
        if not normalized_extensions:
            return []

    file_paths: set[str] = set()
    visited_dirs: set[str] = set()

    def walk_dir(root_dir: str) -> None:
        normalized_root = os.path.normcase(os.path.abspath(root_dir))
        if normalized_root in visited_dirs:
            return
        visited_dirs.add(normalized_root)

        try:
            with os.scandir(root_dir) as entries:
                for entry in entries:
                    entry_path = os.path.abspath(entry.path)
                    try:
                        if entry.is_dir(follow_symlinks=False):
                            if _is_excluded_dir(entry.name):
                                continue
                            walk_dir(entry_path)
                        elif entry.is_file(follow_symlinks=False):
                            if _is_temp_file(entry.name):
                                continue
                            _, ext = os.path.splitext(entry.name)
                            if normalized_extensions is None or ext.lower() in normalized_extensions:
                                file_paths.add(entry_path)
                    except (PermissionError, FileNotFoundError, OSError):
                        pass
        except (PermissionError, FileNotFoundError, NotADirectoryError, OSError):
            pass

    for raw_path in paths:
        if not raw_path:
            continue

        abs_path = os.path.abspath(raw_path)
        if os.path.isfile(abs_path):
            _, ext = os.path.splitext(abs_path)
            if normalized_extensions is None or ext.lower() in normalized_extensions:
                file_paths.add(abs_path)
            continue

        if os.path.isdir(abs_path):
            base_name = os.path.basename(abs_path.rstrip("\\/")).lower()
            if base_name and _is_excluded_dir(base_name):
                continue
            walk_dir(abs_path)

    return sorted(file_paths)


# ──────────────────────────────────────────────
# 텍스트 추출
# ──────────────────────────────────────────────

def extract_text(filepath: str) -> list[tuple[str, str]]:
    """파일 확장자별 텍스트를 추출해 (위치, 텍스트) 목록으로 반환한다."""
    text_items, _ = _extract_text_with_status(filepath)
    return text_items


# ──────────────────────────────────────────────
# 빠른 사전 필터링
# ──────────────────────────────────────────────

def quick_check(filepath: str, keywords: list[str]) -> bool:
    """바이너리 레벨에서 키워드 포함 여부를 사전 확인한다."""
    normalized_keywords = _normalize_keyword_list(keywords)
    if not normalized_keywords:
        return False

    try:
        file_size = os.path.getsize(filepath)
        if file_size > QUICK_CHECK_MAX_BYTES:
            return False
        if file_size == 0:
            return False
    except Exception:
        return True

    patterns = _build_quick_check_patterns(normalized_keywords)
    if not patterns:
        return False

    max_pattern_len = max(len(p) for p in patterns)
    overlap_len = max(0, max_pattern_len - 1)

    try:
        with open(filepath, "rb") as file:
            tail = b""
            while True:
                chunk = file.read(1024 * 1024)
                if not chunk:
                    break
                buffer = tail + chunk
                for pattern in patterns:
                    if pattern in buffer:
                        return True
                tail = buffer[-overlap_len:] if overlap_len > 0 else b""
    except Exception:
        return True

    return False


# ──────────────────────────────────────────────
# 키워드 검색
# ──────────────────────────────────────────────

def search_file(filepath: str, keywords: list[str]) -> list[dict[str, str]]:
    """1단계 quick_check 후 통과 파일만 정밀 추출/검색해 결과를 반환한다."""
    normalized_keywords = _normalize_keyword_list(keywords)
    skipped = False
    failed = False

    try:
        if not normalized_keywords:
            skipped = True
            return []

        results: list[dict[str, str]] = []
        filename = os.path.basename(filepath)
        lowered_filename = filename.lower()
        found_in_filename: set[str] = set()

        for keyword in normalized_keywords:
            lowered_keyword = keyword.lower()
            if lowered_keyword and lowered_keyword in lowered_filename:
                found_in_filename.add(lowered_keyword)
                results.append(
                    {
                        "keyword": keyword,
                        "file": filepath,
                        "location": "파일명",
                        "context": filename,
                    }
                )

        remaining_keywords = [
            kw for kw in normalized_keywords if kw.lower() not in found_in_filename
        ]
        if not remaining_keywords:
            return results

        if not quick_check(filepath, remaining_keywords):
            skipped = True
            return results

        text_items, extract_failed = _extract_text_with_status(filepath)
        if extract_failed:
            failed = True
            return results

        body_matches = search_keywords(text_items, remaining_keywords)
        for item in body_matches:
            results.append(
                {
                    "keyword": str(item.get("keyword", "")),
                    "file": filepath,
                    "location": str(item.get("location", "")),
                    "context": str(item.get("context", "")),
                }
            )
        return results
    except Exception:
        failed = True
        return []
    finally:
        _set_search_file_meta(filepath, skipped=skipped, failed=failed)


def search_file_by_name(filepath: str, keywords: list[str]) -> list[dict[str, str]]:
    """파일명에서만 키워드를 검색해 키워드당 최대 1건 결과를 반환한다."""
    normalized_keywords = _normalize_keyword_list(keywords)
    if not normalized_keywords:
        return []

    filename = os.path.basename(filepath)
    lowered_filename = filename.lower()
    matches: list[dict[str, str]] = []

    for keyword in normalized_keywords:
        lowered_keyword = keyword.lower()
        if lowered_keyword and lowered_keyword in lowered_filename:
            matches.append(
                {
                    "keyword": keyword,
                    "file": filepath,
                    "location": "파일명",
                    "context": filename,
                }
            )

    return matches


def search_keywords(
    text_items: Iterable[tuple[str, str]],
    keywords: list[str],
) -> list[dict[str, str]]:
    """텍스트 목록에서 키워드를 검색해 위치와 문맥 정보를 반환한다."""
    normalized_keywords = _normalize_keyword_list(keywords)
    if not normalized_keywords:
        return []

    results: list[dict[str, str]] = []
    remaining: dict[str, str] = {kw.lower(): kw for kw in normalized_keywords}
    iterator = iter(text_items)

    try:
        for location, text in iterator:
            if not remaining:
                break
            if not text:
                continue

            lowered_text = text.lower()

            for lowered_keyword, keyword in list(remaining.items()):
                found_idx = lowered_text.find(lowered_keyword)
                if found_idx < 0:
                    continue

                context_start = max(0, found_idx - 50)
                context_end = min(len(text), found_idx + len(keyword) + 50)
                context = text[context_start:context_end].strip()

                results.append(
                    {
                        "keyword": keyword,
                        "location": location,
                        "context": context,
                    }
                )
                remaining.pop(lowered_keyword, None)

            if not remaining:
                break
    finally:
        close = getattr(iterator, "close", None)
        if callable(close):
            close()

    return results


def consume_search_file_meta(filepath: str) -> dict[str, bool]:
    """search_file 실행 메타(스킵/실패)를 반환하고 내부 저장소에서 제거한다."""
    normalized = os.path.normcase(os.path.abspath(filepath))
    with _SEARCH_FILE_META_LOCK:
        skipped, failed = _SEARCH_FILE_META.pop(normalized, (False, False))
    return {"skipped": skipped, "failed": failed}


def clear_all_search_file_meta() -> None:
    """내부 메타 저장소를 완전히 비운다."""
    with _SEARCH_FILE_META_LOCK:
        _SEARCH_FILE_META.clear()


# ──────────────────────────────────────────────
# 리포트 저장
# ──────────────────────────────────────────────

def save_report(results: list[dict[str, Any]], output_path: str) -> None:
    """검색 결과를 Excel 리포트 파일로 저장한다."""
    if openpyxl is None:
        return

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "검색결과"

    worksheet.append(["키워드", "파일경로", "위치", "해당문장", "검색일시"])
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row in results:
        keyword = str(row.get("keyword", "") or "")
        file_path = str(row.get("file_path", row.get("filepath", "")) or "")
        location = str(row.get("location", "") or "")
        context = str(row.get("context", "") or "")
        searched_at = str(row.get("searched_at", now) or now)
        worksheet.append([keyword, file_path, location, context, searched_at])

    try:
        parent = os.path.dirname(os.path.abspath(output_path))
        if parent:
            os.makedirs(parent, exist_ok=True)
        workbook.save(output_path)
    except OSError:
        pass
    finally:
        workbook.close()


# ──────────────────────────────────────────────
# 내부 헬퍼 — 텍스트 추출
# ──────────────────────────────────────────────

def _extract_text_with_status(filepath: str) -> tuple[list[tuple[str, str]], bool]:
    try:
        file_size = os.path.getsize(filepath)
        if file_size > MAX_EXTRACT_FILE_SIZE:
            return [], True
    except OSError:
        return [], True

    _, ext = os.path.splitext(filepath)
    ext = ext.lower()

    extractors: dict[str, Callable[[str], list[tuple[str, str]]]] = {
        ".pdf": _extract_pdf,
        ".xlsx": _extract_xlsx,
        ".xls": _extract_xls,
        ".docx": _extract_docx,
        ".doc": _extract_doc,
        ".hwp": _extract_hwp,
        ".hwpx": _extract_hwpx,
        ".csv": _extract_csv,
        ".txt": _extract_txt,
    }

    extractor = extractors.get(ext)
    if extractor is None:
        return [], False

    try:
        return extractor(filepath), False
    except Exception:
        return [], True


def _extract_pdf(filepath: str) -> list[tuple[str, str]]:
    items: list[tuple[str, str]] = []

    if fitz is not None:
        try:
            with fitz.open(filepath) as document:
                page_count = min(len(document), MAX_PDF_PAGES)
                for page_num in range(1, page_count + 1):
                    page = document[page_num - 1]
                    raw_text = page.get_text("text") or ""
                    for line_num, line in enumerate(raw_text.splitlines(), start=1):
                        cleaned = line.strip()
                        if cleaned:
                            items.append((f"P{page_num} L{line_num}", cleaned))
            return items
        except Exception:
            items.clear()

    if pdfplumber is None:
        raise RuntimeError("PDF 추출 라이브러리(fitz/pdfplumber) 없음")

    with pdfplumber.open(filepath) as pdf:
        page_count = min(len(pdf.pages), MAX_PDF_PAGES)
        for page_num in range(1, page_count + 1):
            page = pdf.pages[page_num - 1]
            raw_text = page.extract_text() or ""
            for line_num, line in enumerate(raw_text.splitlines(), start=1):
                cleaned = line.strip()
                if cleaned:
                    items.append((f"P{page_num} L{line_num}", cleaned))

    return items


def _extract_xlsx(filepath: str) -> list[tuple[str, str]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl 미설치")

    items: list[tuple[str, str]] = []
    workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    text = str(value).strip()
                    if text:
                        items.append((f"{sheet.title} {cell.coordinate}", text))
    finally:
        workbook.close()

    return items


def _extract_xls(filepath: str) -> list[tuple[str, str]]:
    if xlrd is None:
        raise RuntimeError("xlrd 미설치")

    items: list[tuple[str, str]] = []
    workbook = xlrd.open_workbook(filepath, on_demand=True, logfile=io.StringIO())

    try:
        for sheet in workbook.sheets():
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    value = sheet.cell_value(row_idx, col_idx)
                    if value in (None, ""):
                        continue
                    text = str(value).strip()
                    if text:
                        items.append(
                            (f"{sheet.name} R{row_idx + 1}C{col_idx + 1}", text)
                        )
    finally:
        try:
            workbook.release_resources()
        except Exception:
            pass

    return items


def _extract_docx(filepath: str) -> list[tuple[str, str]]:
    if Document is None:
        raise RuntimeError("python-docx 미설치")

    items: list[tuple[str, str]] = []
    document = Document(filepath)

    for idx, paragraph in enumerate(document.paragraphs, start=1):
        text = paragraph.text.strip()
        if text:
            items.append((f"문단{idx}", text))

    return items


def _extract_doc(filepath: str) -> list[tuple[str, str]]:
    if olefile is None:
        raise RuntimeError("olefile 미설치")

    items: list[tuple[str, str]] = []

    with olefile.OleFileIO(filepath) as ole:
        if not ole.exists("WordDocument"):
            return []

        try:
            raw = ole.openstream("WordDocument").read()
        except OSError:
            return []

        if len(raw) > BINARY_FALLBACK_MAX_BYTES:
            raw = raw[:BINARY_FALLBACK_MAX_BYTES]

        candidates = _extract_doc_text_candidates(raw)
        for line_no, text in enumerate(candidates, start=1):
            items.append((f"DOC L{line_no}", text))

    return items


def _extract_hwp(filepath: str) -> list[tuple[str, str]]:
    if olefile is not None:
        try:
            if olefile.isOleFile(filepath):
                items = _extract_hwp_ole(filepath)
                if items:
                    return items
        except Exception:
            pass

    try:
        if zipfile.is_zipfile(filepath):
            items = _extract_hwpx(filepath)
            if items:
                return items
    except Exception:
        pass

    try:
        items = _extract_hwpml(filepath)
        if items:
            return items
    except Exception:
        pass

    try:
        return _extract_hwp_binary_fallback(filepath)
    except Exception:
        pass

    return []


def _extract_hwp_ole(filepath: str) -> list[tuple[str, str]]:
    items: list[tuple[str, str]] = []

    with olefile.OleFileIO(filepath) as ole:
        compressed = _is_hwp_compressed(ole)
        streams = _iter_hwp_section_streams(ole)

        line_no = 1
        for stream_name in streams:
            try:
                raw = ole.openstream(stream_name).read()
            except OSError:
                continue

            decompressed = _decompress_hwp_stream(raw, compressed)
            if not decompressed:
                continue

            paragraphs = _parse_hwp_para_text(decompressed)
            for text in paragraphs:
                items.append((f"L{line_no}", text))
                line_no += 1

    return items


def _extract_hwp_binary_fallback(filepath: str) -> list[tuple[str, str]]:
    items: list[tuple[str, str]] = []
    seen: set[str] = set()

    with open(filepath, "rb") as f:
        raw = f.read(BINARY_FALLBACK_MAX_BYTES)

    line_no = 1
    for encoding in ("utf-16le", "cp949", "utf-8"):
        try:
            decoded = raw.decode(encoding, errors="ignore")
        except Exception:
            continue

        for token in re.split(r"[\r\n\t\x00]+", decoded):
            cleaned = _clean_text(token)
            if len(cleaned) < 2:
                continue
            if not re.search(r"[가-힣]", cleaned):
                continue
            lowered = cleaned.lower()
            if lowered in seen:
                continue
            seen.add(lowered)
            items.append((f"L{line_no}", cleaned))
            line_no += 1

    return items


def _extract_hwpml(filepath: str) -> list[tuple[str, str]]:
    items: list[tuple[str, str]] = []
    line_no = 1

    for encoding in ("utf-8", "cp949", "euc-kr"):
        try:
            with open(filepath, "r", encoding=encoding) as f:
                content = f.read()
            break
        except (UnicodeDecodeError, OSError):
            continue
    else:
        return []

    if not content.strip().startswith("<?xml") and "<HWPML" not in content[:500]:
        return []

    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        return _extract_hwpml_regex(content)

    for elem in root.iter():
        tag = elem.tag
        if isinstance(tag, str) and "}" in tag:
            tag = tag.split("}", 1)[1]

        if tag.upper() in ("CHAR", "T", "TEXT"):
            if elem.text:
                cleaned = _clean_text(elem.text)
                if cleaned:
                    items.append((f"L{line_no}", cleaned))
                    line_no += 1

    if not items:
        for raw_text in root.itertext():
            cleaned = _clean_text(raw_text)
            if cleaned and len(cleaned) >= 2 and re.search(r"[가-힣]", cleaned):
                items.append((f"L{line_no}", cleaned))
                line_no += 1

    return items


def _extract_hwpml_regex(content: str) -> list[tuple[str, str]]:
    items: list[tuple[str, str]] = []
    line_no = 1
    seen: set[str] = set()

    for match in re.finditer(r"<(?:CHAR|T|TEXT)[^>]*>([^<]+)</", content):
        cleaned = _clean_text(match.group(1))
        if cleaned and len(cleaned) >= 2 and re.search(r"[가-힣]", cleaned):
            lowered = cleaned.lower()
            if lowered not in seen:
                seen.add(lowered)
                items.append((f"L{line_no}", cleaned))
                line_no += 1

    if not items:
        for match in re.finditer(r">([^<]+)<", content):
            cleaned = _clean_text(match.group(1))
            if cleaned and len(cleaned) >= 2 and re.search(r"[가-힣]", cleaned):
                lowered = cleaned.lower()
                if lowered not in seen:
                    seen.add(lowered)
                    items.append((f"L{line_no}", cleaned))
                    line_no += 1

    return items


def _extract_hwpx(filepath: str) -> list[tuple[str, str]]:
    items: list[tuple[str, str]] = []
    line_no = 1

    with zipfile.ZipFile(filepath, "r") as archive:
        xml_files = [
            name
            for name in archive.namelist()
            if name.lower().endswith(".xml") and "section" in name.lower()
        ]
        if not xml_files:
            xml_files = [
                name for name in archive.namelist() if name.lower().endswith(".xml")
            ]

        for xml_name in sorted(xml_files):
            try:
                xml_bytes = archive.read(xml_name)
                root = ET.fromstring(xml_bytes)
            except (KeyError, zipfile.BadZipFile, ET.ParseError):
                continue

            for raw_text in root.itertext():
                cleaned = _clean_text(raw_text)
                if cleaned:
                    items.append((f"L{line_no}", cleaned))
                    line_no += 1

    return items


def _extract_csv(filepath: str) -> list[tuple[str, str]]:
    items: list[tuple[str, str]] = []

    for encoding in TXT_FALLBACK_ENCODINGS:
        try:
            with open(filepath, "r", encoding=encoding, newline="") as file:
                reader = csv.reader(file)
                for row_idx, row in enumerate(reader, start=1):
                    for col_idx, value in enumerate(row, start=1):
                        cleaned = value.strip()
                        if cleaned:
                            items.append((f"R{row_idx}C{col_idx}", cleaned))
            return items
        except UnicodeDecodeError:
            items.clear()
            continue

    return items


def _extract_txt(filepath: str) -> list[tuple[str, str]]:
    for encoding in TXT_FALLBACK_ENCODINGS:
        try:
            with open(filepath, "r", encoding=encoding) as file:
                lines = file.readlines()
            return [
                (f"L{line_no}", line.strip())
                for line_no, line in enumerate(lines, start=1)
                if line.strip()
            ]
        except UnicodeDecodeError:
            continue

    return []


# ──────────────────────────────────────────────
# 내부 헬퍼 — 유틸리티
# ──────────────────────────────────────────────

def _normalize_extensions(extensions: set[str]) -> set[str]:
    normalized: set[str] = set()
    for extension in extensions:
        if not extension:
            continue
        cleaned = extension.strip().lower()
        if not cleaned:
            continue
        if not cleaned.startswith("."):
            cleaned = f".{cleaned}"
        normalized.add(cleaned)
    return normalized


def _build_quick_check_patterns(keywords: list[str]) -> list[bytes]:
    patterns: list[bytes] = []
    seen: set[bytes] = set()

    for keyword in keywords:
        has_alpha = any(c.isascii() and c.isalpha() for c in keyword)
        if has_alpha:
            text_variants = [
                keyword, keyword.lower(), keyword.upper(), keyword.title(),
            ]
        else:
            text_variants = [keyword]

        for encoding in QUICK_CHECK_ENCODINGS:
            for text in text_variants:
                try:
                    encoded = text.encode(encoding)
                except (UnicodeEncodeError, UnicodeDecodeError):
                    continue

                if not encoded or encoded in seen:
                    continue
                seen.add(encoded)
                patterns.append(encoded)

    patterns.sort(key=len)
    return patterns


def _set_search_file_meta(filepath: str, skipped: bool, failed: bool) -> None:
    normalized = os.path.normcase(os.path.abspath(filepath))
    with _SEARCH_FILE_META_LOCK:
        _SEARCH_FILE_META[normalized] = (skipped, failed)


def _normalize_keyword_list(keywords: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()

    for keyword in keywords:
        if not isinstance(keyword, str):
            continue
        cleaned = keyword.strip()
        if not cleaned:
            continue
        lowered = cleaned.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        normalized.append(cleaned)

    return normalized


def _is_excluded_dir(dir_name: str) -> bool:
    return dir_name.lower() in EXCLUDED_DIR_NAMES


def _is_temp_file(filename: str) -> bool:
    """Office 임시 파일(~$...)만 제외한다."""
    return filename.startswith("~$")


def _is_hwp_compressed(ole: Any) -> bool:
    if not ole.exists("FileHeader"):
        return False

    file_header = ole.openstream("FileHeader").read()
    if len(file_header) < 40:
        return False

    flags = int.from_bytes(file_header[36:40], byteorder="little", signed=False)
    return bool(flags & 0x01)


def _iter_hwp_section_streams(ole: Any) -> list[str]:
    stream_names: list[str] = []
    for path_parts in ole.listdir(streams=True, storages=False):
        if (
            len(path_parts) >= 2
            and path_parts[0] == "BodyText"
            and path_parts[1].startswith("Section")
        ):
            stream_names.append("/".join(path_parts))
    return stream_names


def _decompress_hwp_stream(data: bytes, compressed: bool) -> bytes:
    if not compressed:
        return data

    try:
        return zlib.decompress(data, -15)
    except zlib.error:
        try:
            return zlib.decompress(data)
        except zlib.error:
            return b""


def _parse_hwp_para_text(data: bytes) -> list[str]:
    paragraphs: list[str] = []
    offset = 0
    data_len = len(data)

    while offset + 4 <= data_len:
        header = int.from_bytes(
            data[offset : offset + 4], byteorder="little", signed=False
        )
        offset += 4

        record_type = header & 0x3FF
        record_size = (header >> 20) & 0xFFF

        if record_size == 0xFFF:
            if offset + 4 > data_len:
                break
            record_size = int.from_bytes(
                data[offset : offset + 4], byteorder="little", signed=False
            )
            offset += 4

        end_offset = offset + record_size
        if end_offset > data_len:
            break

        payload = data[offset:end_offset]
        offset = end_offset

        if record_type != 67 or not payload:
            continue

        decoded = payload.decode("utf-16le", errors="ignore")
        cleaned = _clean_text(decoded)
        if cleaned:
            paragraphs.append(cleaned)

    return paragraphs


def _clean_text(text: str) -> str:
    cleaned = text.replace("\x00", " ")
    cleaned = re.sub(r"[\x01-\x08\x0B\x0C\x0E-\x1F]", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def _extract_doc_text_candidates(raw: bytes) -> list[str]:
    candidates: list[str] = []
    seen: set[str] = set()

    def collect(decoded_text: str) -> None:
        for token in re.split(r"[\r\n\t]+", decoded_text):
            cleaned = _clean_text(token)
            if len(cleaned) < 2:
                continue
            if not re.search(r"[0-9A-Za-z가-힣]", cleaned):
                continue
            lowered = cleaned.lower()
            if lowered in seen:
                continue
            seen.add(lowered)
            candidates.append(cleaned)

    collect(raw.decode("utf-16le", errors="ignore"))
    collect(raw.decode("cp949", errors="ignore"))
    collect(raw.decode("latin-1", errors="ignore"))

    return candidates
